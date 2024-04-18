import sys
import os
import openpyxl
import tkinter as tk
import warnings
from tkinter import filedialog, messagebox, ttk
from tabulate import tabulate

# UserWarning ausblenden 
warnings.simplefilter(action='ignore', category=UserWarning)

# Initialize Tkinter
root = tk.Tk()
root.title("Excel-to-Mardown-Converter")

# Script beenden wenn Fenster geschlossen wird
root.bind("<Destroy>", lambda e: sys.exit())

# Funktion zum Öffnen des Dateiauswahldialogs
def open_file_dialog():
    file_path = filedialog.askopenfilename(
        parent=root,
        title="Bitte wählen Sie das Excel-Dokument aus.",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if file_path:
        process_workbook(file_path)

# Funktion zum Verarbeiten der ausgewählten Excel-Datei
def process_workbook(file_path):
    try:
        # Excel-Datei mit openpyxl laden
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        messagebox.showinfo("Fehler", f"Fehler beim Laden der Excel-Datei: {str(e)}")
        return
    
    # List of available formats
    formats = ["pipe", "yaml", "xml", "html", "csv", "json", "plain", "jira", "mediawiki"]

    # Create a StringVar for the combobox
    selected_format = tk.StringVar()


    # Label erstellen
    tk.Label(root, text= "Excel-Arbeitsblätter zum Exportieren auswählen:").pack(anchor='w')

    # Liste der verfügbaren Arbeitsblätter
    available_sheets = workbook.sheetnames

    # Check ob Arbeitsmappen vorhanden sind
    if not available_sheets:
        messagebox.showinfo("Fehler", "Die ausgewählte Excel-Datei enthält keine Arbeitsblätter.")
        return

    # Listbox für die Auswahl der Arbeitsblätter
    listbox = tk.Listbox(root, selectmode='multiple', exportselection=0)
    listbox.pack(side='top', fill='both', expand=True)

    # Arbeitsblätter zur Listbox hinzufügen
    for sheet in available_sheets:
        listbox.insert(tk.END, sheet)

    # Initialize Checkbox
    export_visible = tk.BooleanVar(value=True)

    # Create Checkbox
    tk.Checkbutton(root, text="Ausgeblendete Zeilen/Spalten beim Export nicht einbeziehen", variable=export_visible).pack(anchor='w')

    # Label for the combobox
    tk.Label(root, text="Wähle gewünschtes Format:").pack(anchor='w')

    # Create the combobox
    combobox = ttk.Combobox(root, textvariable=selected_format, state='readonly')
    combobox['values'] = formats
    combobox.current(0)  # set initial selection
    combobox.pack(anchor='w')

    # Funktionen für die Buttons
    def select_all():
        listbox.select_set(0, tk.END)

    def deselect_all():
        listbox.selection_clear(0, tk.END)

    def cancel():
        root.destroy()

    from openpyxl.utils import column_index_from_string

    def export_sheets():
        selected_indices = listbox.curselection()
        selected_sheets = [listbox.get(i) for i in selected_indices]

        # Check ob Arbeitsmappen ausgewählt wurden
        if not selected_sheets:
            messagebox.showinfo("Information", "Bitte wähle Arbeitsmappen aus")
            return

        for sheet_name in selected_sheets:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Collect data
                data = []

                for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
                    # Prüft ob Spalte versteckt ist - wenn Checkbox aktiviert
                    if export_visible.get() and sheet.row_dimensions[row[0].row].hidden:
                        continue
                    row_data = []
                    for cell in row:
                        # Prüft ob Zeile versteckt ist - wenn Checkbox aktiviert
                        column_dimension = sheet.column_dimensions.get(cell.column_letter)
                        if not export_visible.get() or (column_dimension is not None and not column_dimension.hidden):
                            row_data.append(cell.value)
                    # Wenn Spalte nicht leer, in Data übertragen
                    if row_data:
                        data.append(row_data)

                # Daten in gewähltes Format konvertieren
                format_table = tabulate(data, tablefmt=selected_format.get(), headers="firstrow", showindex=False)
 
                # Dictionary mapping formats to file extensions
                format_extensions = {"pipe": ".md", "plain": ".txt", "jira": ".txt", "mediawiki": ".txt", "csv": ".csv", "json": ".json", "yaml": ".yaml", "xml": ".xml", "html": ".html"}

                # Get the file extension for the selected format
                file_extension = format_extensions[selected_format.get()]

                # Pfad für die Markdown-Datei setzen
                output_file_path = f"{file_path.rsplit('.', 1)[0]}_{sheet_name}{file_extension}"

                # Check ob Dateiname bereits existiert
                if os.path.isfile(output_file_path):
                    overwrite = messagebox.askyesno("Datei existiert bereits", f"Die Datei {output_file_path} existiert bereits. Überschreiben?")
                    if not overwrite:
                        return

                # Markdown-Tabelle in eine Datei schreiben
                try:
                    with open(output_file_path, "w", encoding="utf-8") as f:
                        f.write(format_table)
                except Exception as e:
                    messagebox.showinfo("Fehler", f"Fehler beim Schreiben der Datei {output_file_path}: {str(e)}")
                    return

        # Erfolgsmeldung anzeigen
        messagebox.showinfo("Erfolg", "File(s) erfolgreich erstellt.")
        root.quit()

    # Buttons
    button_select_all = tk.Button(root, text="Alle auswählen", command=select_all)
    button_select_all.pack(side='left', fill='x', expand=True)

    button_deselect_all = tk.Button(root, text="Auswahl aufheben", command=deselect_all)
    button_deselect_all.pack(side='left', fill='x', expand=True)

    button_export = tk.Button(root, text="Exportieren", command=export_sheets)
    button_export.pack(side='left', fill='x', expand=True)

    button_cancel = tk.Button(root, text="Abbrechen", command=cancel)
    button_cancel.pack(side='left', fill='x', expand=True)

# Öffne den Dateiauswahldialog
open_file_dialog()

# Tkinter Event-Schleife starten
root.mainloop()
