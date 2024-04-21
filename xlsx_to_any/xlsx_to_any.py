# Standard Library Imports
import sys
import os
import csv
import json
import xml.etree.ElementTree as ET
import warnings
from tkinter import filedialog, messagebox, ttk
from xml.etree.ElementTree import Element, SubElement, ElementTree

# Third-party Imports
import openpyxl
import tkinter as tk
from tabulate import tabulate
import yaml
import pandas as pd

def main():
    # Suppress terminal note for input validation
    warnings.simplefilter(action='ignore', category=UserWarning)

    # Initialize Tkinter
    root = tk.Tk()
    root.title("Excel-to-Any-Converter")

    # Exit script if window is closed
    root.bind("<Destroy>", lambda e: sys.exit())

    # Function to open file dialog
    def open_file_dialog():
        file_path = filedialog.askopenfilename(
            parent=root,
            title="Please select the Excel document.",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            process_workbook(file_path)
        else:
            sys.exit()
            
    # Function to process the selected Excel file
    def process_workbook(file_path):
        try:
            # Load Excel file with openpyxl
            workbook = openpyxl.load_workbook(file_path, data_only=True)
                
        except Exception as e:
            messagebox.showinfo("Error", f"Error loading the Excel file: {str(e)}")
            return
        
        # List of available formats
        formats = ["pipe", "yaml", "xml", "html", "csv", "json", "plain", "jira", "mediawiki"]

        # Create StringVar for combobox
        selected_format = tk.StringVar()

        # Create Label
        tk.Label(root, text= "Select Excel sheets to export:").pack(anchor='w')

        # List of available sheets
        available_sheets = workbook.sheetnames

        # Check if worksheets are present
        if not available_sheets:
            messagebox.showinfo("Error", "The selected Excel file contains no sheets.")
            return

        # Listbox for selecting sheets
        listbox = tk.Listbox(root, selectmode='multiple', exportselection=0)
        listbox.pack(side='top', fill='both', expand=True)

        # Add worksheets to listbox
        for sheet in available_sheets:
            listbox.insert(tk.END, sheet)

        # Initialize Checkbox
        export_visible = tk.BooleanVar(value=True)

        # Create Checkbox
        tk.Checkbutton(root, text="Exclude hidden rows/columns from export", variable=export_visible).pack(anchor='w')

        # Label for combobox
        tk.Label(root, text="Choose desired format:").pack(anchor='w')

        # Create combobox
        combobox = ttk.Combobox(root, textvariable=selected_format, state='readonly')
        combobox['values'] = formats
        combobox.current(0)  # set initial selection
        combobox.pack(anchor='w')

        # Functions for buttons
        def select_all():
            listbox.select_set(0, tk.END)

        def deselect_all():
            listbox.selection_clear(0, tk.END)

        def cancel():
            root.destroy()

        from openpyxl.utils import column_index_from_string
        
        # Export Sheet Function
        def export_sheets():
            selected_indices = listbox.curselection()
            selected_sheets = [listbox.get(i) for i in selected_indices]

            # Check if worksheets are selected
            if not selected_sheets:
                messagebox.showinfo("No sheets selected", "Please select at least one sheet.")
                return

            # Dictionary Mapping formats to file extensions
            format_extensions = {"pipe": ".md", "plain": ".txt", "jira": ".txt", "mediawiki": ".txt", "csv": ".csv", "json": ".json", "yaml": ".yaml", "xml": ".xml", "html": ".html"}

            for sheet_name in selected_sheets:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Unmerge any merged cells
                    for merged_cell_range in list(sheet.merged_cells.ranges):
                        sheet.unmerge_cells(str(merged_cell_range))

                    # Collect data
                    data = []

                    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
                        # Check if column is hidden - if checkbox is enabled
                        if export_visible.get() and sheet.row_dimensions[row[0].row].hidden:
                            continue
                        row_data = []
                        for cell in row:
                            # Check if row is hidden - if checkbox is enabled
                            column_dimension = sheet.column_dimensions.get(cell.column_letter)
                            if not export_visible.get() or (column_dimension is not None and not column_dimension.hidden):
                                row_data.append(cell.value)
                        # If column is not empty, transfer to data
                        if row_data:
                            data.append(row_data)

                    # Convert data to selected format (XML; JSON; YAML; TABULATE)
                    if selected_format.get() == "xml":
                        output_file_path = f"{file_path.rsplit('.', 1)[0]}_{sheet_name}.xml"
                        success = write_xml_file(data, output_file_path)

                    elif selected_format.get() == "json":
                        output_file_path = f"{file_path.rsplit('.', 1)[0]}_{sheet_name}.json"
                        success = write_json_file(data, output_file_path)

                    elif selected_format.get() == "yaml":
                        output_file_path = f"{file_path.rsplit('.', 1)[0]}_{sheet_name}.yaml"
                        success = write_yaml_file(data, output_file_path)
                        
                    else:
                        output = tabulate(data, tablefmt=selected_format.get())
                        output_file_path = f"{file_path.rsplit('.', 1)[0]}_{sheet_name}{format_extensions[selected_format.get()]}"
                        success = write_other_formats(output, output_file_path)

            if success:
                messagebox.showinfo("Success", "Data exported successfully.")
        
        # XML Function
        def write_xml_file(data, output_file_path):
            # Check if file already exists
            if os.path.isfile(output_file_path):
                overwrite = messagebox.askyesno("File already exists", f"The file {output_file_path} already exists. Overwrite?")
                if not overwrite:
                    return False

            root = ET.Element("root")

            for row in data:
                row_element = ET.SubElement(root, "row")
                for cell in row:
                    cell_element = ET.SubElement(row_element, "cell")
                    cell_element.text = str(cell)

            tree = ET.ElementTree(root)

            try:
                tree.write(output_file_path)
            except Exception as e:
                messagebox.showinfo("Error", f"Error writing the file {output_file_path}: {str(e)}")
                return False

            return True

        # JSON Function
        def write_json_file(data, output_file_path):
            # Check if file already exists
            if os.path.isfile(output_file_path):
                overwrite = messagebox.askyesno("File already exists", f"The file {output_file_path} already exists. Overwrite?")
                if not overwrite:
                    return False

            try:
                with open(output_file_path, "w") as f:
                    json.dump(data, f)
            except Exception as e:
                messagebox.showinfo("Error", f"Error writing the file {output_file_path}: {str(e)}")
                return False
            
            return True

        # YAML Function
        def write_yaml_file(data, output_file_path):
            # Check if file already exists
            if os.path.isfile(output_file_path):
                overwrite = messagebox.askyesno("File already exists", f"The file {output_file_path} already exists. Overwrite?")
                if not overwrite:
                    return False

            try:
                with open(output_file_path, "w") as f:
                    yaml.dump(data, f)
            except Exception as e:
                messagebox.showinfo("Error", f"Error writing the file {output_file_path}: {str(e)}")
                return False

            return True

        # Supported formats function
        def write_other_formats(output, output_file_path):
            # Check if file already exists
            if os.path.isfile(output_file_path):
                overwrite = messagebox.askyesno("File already exists", f"The file {output_file_path} already exists. Overwrite?")
                if not overwrite:
                    return False

            try:
                with open(output_file_path, "w") as f:
                    f.write(output)
            except Exception as e:
                messagebox.showinfo("Error", f"Error writing the file {output_file_path}: {str(e)}")
                return False

            return True     

        # Buttons
        button_select_all = tk.Button(root, text="Select All", command=select_all)
        button_select_all.pack(side='left', fill='x', expand=True)

        button_deselect_all = tk.Button(root, text="Deselect All", command=deselect_all)
        button_deselect_all.pack(side='left', fill='x', expand=True)

        button_export = tk.Button(root, text="Export", command=export_sheets)
        button_export.pack(side='left', fill='x', expand=True)

        button_cancel = tk.Button(root, text="Cancel", command=cancel)
        button_cancel.pack(side='left', fill='x', expand=True)

    # Open file dialog
    open_file_dialog()

    # Start Tkinter event loop
    root.mainloop()
